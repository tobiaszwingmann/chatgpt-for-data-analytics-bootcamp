from pathlib import Path
import openpyxl
import pandas as pd

SOURCE_XLSX = Path("consolidated_customer_report.xlsx")
OUTPUT_CSV = Path("customer_quarterly_tidy.csv")


def extract_section(ws, section_name, header_row):
    """Return raw customer rows for one report section."""
    rows = []
    r = header_row + 1

    while r <= ws.max_row:
        values = [ws.cell(r, c).value for c in range(1, 15)]
        first_cell = values[0]

        # Stop at the section totals row
        if isinstance(first_cell, str) and first_cell.strip() == "TOTALS":
            break

        # Ignore fully blank rows
        if any(v is not None for v in values):
            rows.append(values)

        r += 1

    return rows


def main():
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=False)
    ws = wb["Sheet1"]

    sections = {
        "National": 5,       # header row
        "International": 238 # header row
    }

    tidy_rows = []

    for section_name, header_row in sections.items():
        raw_rows = extract_section(ws, section_name, header_row)

        for values in raw_rows:
            customer_id = values[0]
            order_q = values[1:5]
            revenue_q = values[5:9]
            lifetime_revenue = values[9]
            country = values[10]
            city = values[11]
            marketing_opt_in = values[12]
            email_domain = values[13]

            for q_index, quarter in enumerate(["Q1", "Q2", "Q3", "Q4"]):
                tidy_rows.append(
                    {
                        "section": section_name,
                        "customer_id": int(customer_id),
                        "quarter": quarter,
                        "order_volume": int(order_q[q_index]),
                        "revenue_gbp": float(revenue_q[q_index]),
                        "lifetime_revenue_gbp": float(lifetime_revenue),
                        "country": country,
                        "city": city,
                        "marketing_opt_in": bool(marketing_opt_in),
                        "email_domain": email_domain,
                    }
                )

    tidy_df = pd.DataFrame(
        tidy_rows,
        columns=[
            "section",
            "customer_id",
            "quarter",
            "order_volume",
            "revenue_gbp",
            "lifetime_revenue_gbp",
            "country",
            "city",
            "marketing_opt_in",
            "email_domain",
        ],
    )

    tidy_df.to_csv(OUTPUT_CSV, index=False)
    print(f"Wrote {len(tidy_df)} rows to {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
